# SPDX-FileCopyrightText: 2026 Bobby Yu
# SPDX-License-Identifier: AGPL-3.0-or-later

"""PM document sidecar.

A long-lived helper process that does the jobs the Rust backend cannot do
natively: convert arbitrary files to Markdown (MarkItDown), embed text locally
with an on-device ONNX model (fastembed), re-rank passages with a local
cross-encoder (fastembed), transcribe voice clips locally (faster-whisper), and
read text + capture metadata out of photos (rapidocr OCR + Pillow EXIF — an
OPTIONAL component the user installs on demand, like the t-SNE reducer).
It is the only part of PM that runs Python, and it is deliberately dumb: it
converts, embeds, scores, transcribes, and analyses images, nothing else.
Chunking, hashing, prefixes, the vault, and the database all live in Rust.

Protocol: newline-delimited JSON over stdio. One request per line, where method is one of
ping, convert, embed, count_tokens, rerank, transcribe, reduce, analyze_image, analyze_spreadsheet:

    {"id": <int>, "method": <name>, "params": {...}}

One response per line:

    {"id": <int>, "ok": true,  "result": {...}}
    {"id": <int>, "ok": false, "error": "message"}

Heavy objects (the MarkItDown instance, the embedding/reranking models, the
Whisper model) are created lazily on first use so `ping` stays instant and
startup is cheap. The long-lived worker runs OFFLINE (PM_SIDECAR_OFFLINE=1): a
model missing from the local cache raises `ModelNotCached`, which the Rust side
turns into a run of this same script with `--fetch` (network allowed) to download
it, then a retry. So the worker that parses untrusted files never needs a socket;
only the short-lived fetcher does (issue #286).

Security: ingested content — including transcribed audio and passages scored by
the reranker — is untrusted data, never instructions. This process only
converts/embeds/scores/transcribes bytes; it never executes file contents.
"""

import ctypes
import json
import logging
import math
import os
import sys
import traceback

# Network posture (issue #286: sidecar OS sandbox). The long-lived WORKER parses untrusted
# file bytes and must run with no outbound sockets; downloading a model is instead the job
# of the short-lived `--fetch` helper, which Rust runs with the network allowed. Rust sets
# PM_SIDECAR_OFFLINE=1 on the worker only. Translate it to Hugging Face's offline flags HERE,
# at import time, before any loader pulls in huggingface_hub (fastembed and faster-whisper
# both fetch through it) — so a cold-cache load fails fast with a catchable error the worker
# can turn into "please fetch this", instead of silently reaching out to the network.
_OFFLINE = os.environ.get("PM_SIDECAR_OFFLINE") == "1"
if _OFFLINE:
    # Hard-assign, NOT setdefault: Rust (via PM_SIDECAR_OFFLINE) is the authority on offline, so an
    # inherited falsy value like HF_HUB_OFFLINE=0 must NOT win. huggingface_hub reads "0" as online
    # and fastembed keys its offline gate off this var, so a stale "0" would let a cold-cache
    # embed/rerank download from inside the untrusted-file worker (the exact socket #286 forbids),
    # with _OFFLINE still True so nothing notices. faster-whisper and (below) fastembed also pass
    # local_files_only=_OFFLINE as belt-and-suspenders, but this env var must be authoritative too.
    os.environ["HF_HUB_OFFLINE"] = "1"
    os.environ["TRANSFORMERS_OFFLINE"] = "1"
    # fastembed logs its download attempts through loguru; offline, those are just noise about a
    # cold-cache miss the worker already turns into a clean "please fetch". Quiet them. Best-effort:
    # loguru is only a fastembed transitive dep, so a partial venv without it must not break import.
    try:
        from loguru import logger

        logger.disable("fastembed")
    except Exception:
        pass

# Model-cache root (issue #286). Rust points every model cache at PM's data dir (runtime/models) via
# PM_MODELS_DIR, so the weights uninstall with the app AND the Windows sidecar sandbox's filesystem
# allow-set is one tidy subtree instead of scattered %TEMP% / ~/.cache paths. The worker and the
# --fetch helper both receive it, so they share one location. Unset (a raw dev run) falls back
# to each library's own default.
_MODELS_DIR = os.environ.get("PM_MODELS_DIR") or None
if _MODELS_DIR:
    # Covers the huggingface_hub / tokenizers cache (the `from_pretrained` fallback path).
    os.environ["HF_HOME"] = os.path.join(_MODELS_DIR, "hf")
    # hf_xet (pulled in by huggingface_hub) does NOT honour HF_HOME: it has its own variable and
    # otherwise writes its chunk cache AND its logs to ~/.cache/huggingface/xet, outside everything
    # PM owns and outside everything the erase visits. Found there on a dev machine, with dated
    # logs. Pinning it keeps the whole model story inside one deletable subtree, which is also what
    # the Windows sandbox's filesystem allow-set assumes.
    os.environ["HF_XET_CACHE"] = os.path.join(_MODELS_DIR, "hf", "xet")


# Resource posture. This is a background helper on someone's own laptop, not a server, and it must
# never be the reason the machine stops feeling responsive. Everything here is set at IMPORT time
# because the native libraries read their thread counts as they load, and every import in this file
# is lazy (inside a handler), so module scope is early enough for all of them.
#
# Left alone, three separate pools each size themselves to the core count: onnxruntime's intra-op
# pool, the OpenBLAS bundled inside numpy, and the rayon pool inside the Rust `tokenizers`
# extension. On a 24-core machine that produced ~94 threads contending for 24 cores while a single
# embedding pass ran — the app "maxing out the CPU" on open. Rust sizes the pool (it has already
# scanned the hardware) and passes PM_SIDECAR_THREADS; a raw dev run with the variable unset
# derives the same figure here so the two paths behave identically.
def _thread_budget():
    """How many threads any one native pool may use. Half the cores, clamped to 2..8.

    The clamp matters more than the ratio: measured on a 24-core box, going from an unbounded pool
    to 8 threads cost ~14% embedding throughput and handed back 16 cores, which is the trade this
    process should always take. The floor of 2 keeps a single-core VM from serializing completely.
    """
    override = os.environ.get("PM_SIDECAR_THREADS")
    if override:
        try:
            return max(1, int(override))
        except ValueError:
            pass
    return max(2, min(8, (os.cpu_count() or 4) // 2))


_THREADS = _thread_budget()

# The env-var half of the budget: numpy's OpenBLAS, any OpenMP runtime under onnxruntime, and
# tokenizers' rayon pool. setdefault, not hard assignment — an operator who deliberately exported
# one of these outranks our default (unlike the offline flags above, where Rust IS the authority).
# onnxruntime's own pool is NOT here: it ignores these and is set per-session via `threads=` in
# get_embedder / get_reranker.
for _var in (
    "OMP_NUM_THREADS",
    "OPENBLAS_NUM_THREADS",
    "MKL_NUM_THREADS",
    "NUMEXPR_NUM_THREADS",
    "RAYON_NUM_THREADS",
):
    os.environ.setdefault(_var, str(_THREADS))
del _var


def _lower_own_priority():
    """Drop below normal scheduling priority, so the OS hands our cores to whatever the user is
    doing the moment they want them.

    This is the piece that makes PM yield without having to watch window focus: a niced process
    still gets the whole machine when nothing else wants it, and loses only under contention —
    which is exactly the desired behaviour and needs no policy of our own. +5 rather than +19: we
    want to lose ties to the user's foreground work, not to be starved behind every daemon.

    Best-effort by design. Lowering priority never requires privilege, but a container with a
    restrictive seccomp profile can still refuse it, and that must not stop the sidecar booting.
    """
    try:
        if hasattr(os, "nice"):
            os.nice(5)
        elif os.name == "nt":
            kernel32 = ctypes.windll.kernel32
            # BELOW_NORMAL_PRIORITY_CLASS — the Windows analogue of a modest nice increment.
            kernel32.SetPriorityClass(kernel32.GetCurrentProcess(), 0x00004000)
    except Exception:
        pass


def _release_free_memory():
    """Return glibc's free lists to the OS. A no-op anywhere else, and harmless when it is.

    Python freeing an object only returns it to malloc, and malloc only returns memory to the
    kernel from the top of an arena — so after a burst of large allocations the process can sit on
    hundreds of MB it will never use again. `malloc_trim` walks the arenas and gives back whole
    free pages. Called after the heavy handlers, where the difference is measurable; calling it
    after `ping` would just be latency.
    """
    try:
        libc = ctypes.CDLL(None)
    except (OSError, TypeError):  # not a POSIX loader (Windows), nothing to trim
        return
    trim = getattr(libc, "malloc_trim", None)
    if trim is None:  # musl and friends: no such symbol
        return
    try:
        trim(0)
    except Exception:
        pass


# Handlers heavy enough to be worth a trim once they return — the ones that run a model or hold a
# whole file in memory. Everything else returns promptly to a small steady state on its own.
_TRIM_AFTER = frozenset(
    {"embed", "rerank", "transcribe", "convert", "analyze_image", "analyze_spreadsheet", "reduce"}
)


# The session posture both ONNX models are built with. fastembed exposes exactly these two knobs
# (`fastembed/common/onnx_model.py`), and between them they are the whole memory story:
#
#   threads               -> intra_op_num_threads / inter_op_num_threads. Unset, onnxruntime takes
#                            one thread per physical core.
#   enable_cpu_mem_arena  -> onnxruntime's CPU arena. Defaults ON, and the arena NEVER returns a
#                            chunk to the OS. One embedding pass at fastembed's default batch of
#                            256 asks for a 256x12x512x512 fp32 attention tensor = exactly 3 GiB,
#                            and the arena then holds it for the life of the process. Measured on
#                            a 24-core box: one embed() call took a fresh interpreter from 0.22 GiB
#                            to 5.31 GiB resident, and neither gc.collect() nor malloc_trim got it
#                            back. With the arena off the same call peaks at 0.47 GiB and settles
#                            to 0.18 GiB. The arena's job is to recycle same-shaped buffers, which
#                            is a real (~20% here) throughput win on a server that never gives
#                            memory back; on a laptop that runs for sixteen hours between restarts
#                            it is the wrong trade.
#
# Bounding the batch (Rust's job) caps the size of any ONE tensor; turning the arena off is what
# makes the memory come back afterwards. Both are needed — with the arena left on, even a batch of
# 32 still settled at 1.22 GiB.
def _session_posture():
    return {"threads": _THREADS, "enable_cpu_mem_arena": False}


def _fastembed_cache_dir():
    """fastembed's cache dir under the shared model root (embed + rerank ONNX weights), or None
    (fastembed's own default) when PM_MODELS_DIR is unset."""
    return os.path.join(_MODELS_DIR, "fastembed") if _MODELS_DIR else None


def _rapidocr_model_dir():
    """rapidocr's detection/recognition model dir under the shared model root, or None (rapidocr's
    own default) when PM_MODELS_DIR is unset.

    rapidocr defaults to `<site-packages>/rapidocr/models` — INSIDE the managed venv. That is the
    one model store PM does not control: rebuilding the venv (or removing the optional OCR
    component, which is a one-click Settings action) throws the weights away, so the next photo
    re-downloads them. Pinning them beside every other model puts the whole model story in one
    subtree — the one the Windows sandbox's filesystem allow-set grants and the one "Remove PM
    data" walks."""
    return os.path.join(_MODELS_DIR, "rapidocr") if _MODELS_DIR else None


# pdfminer (which MarkItDown uses to extract text from PDFs) logs a warning for every glyph
# whose font descriptor has no FontBBox ("None cannot be parsed as 4 floats") — cosmetic noise
# that can flood the console on a single PDF and tells us nothing actionable. Quiet that logger
# to errors; genuine extraction failures still surface.
logging.getLogger("pdfminer").setLevel(logging.ERROR)

# H-1 subset: harden the stdlib XML parser against entity-expansion / XXE before any parser that
# might use it (openpyxl's stdlib fallback when reading an .xlsx). Best-effort — a partial venv
# without defusedxml proceeds rather than breaking ingest. This patches the STDLIB parser only; when
# lxml is present openpyxl may use it instead, but lxml resolves no external entities and caps
# entity expansion by default, so the .xlsx path stays safe either way.
try:
    import defusedxml

    defusedxml.defuse_stdlib()
except Exception:
    pass

# The default embedder + dimension (the English model). PR 2 makes embedding
# model-parameterised: Rust sends the selected model id (and, for a custom model
# not bundled in fastembed, a registration spec) per request, so a vault can use a
# multilingual embedder. These remain the fallback when no model is supplied.
EMBED_MODEL = "BAAI/bge-small-en-v1.5"
EMBED_DIM = 384

# The speech-to-text model for voice input. Fixed for v1, like EMBED_MODEL.
# English-only, ~145 MB, run int8 on CPU — the speed/accuracy sweet spot for
# dictation. Weights download once on first `transcribe`, then it is fully local.
WHISPER_MODEL = "base.en"

_markitdown = None
# Per-model caches, keyed by model id, so one process can serve several models
# (e.g. the English embedder plus a multilingual one). `_registered` tracks the
# custom models already handed to fastembed's `add_custom_model` (calling it twice
# for the same id raises).
_embedders = {}
_tokenizers = {}
_rerankers = {}
_registered = set()
_whisper = None
_ocr_engine = None
_heif_registered = False


class ModelNotCached(Exception):
    """OFFLINE worker reached a model that isn't downloaded yet (issue #286). The main loop
    turns this into a `model_not_cached` reply so Rust runs the network-allowed `--fetch`
    helper and retries the call. Never raised outside offline mode."""


class Unconvertible(Exception):
    """The engine ANSWERED and refused THIS FILE: a verdict about the input, not about the
    engine. The main loop turns it into an `unconvertible` reply, which is the ONLY signal
    `cloud_sync::is_permanently_unindexable` accepts as "skip this item and let the delta
    cursor move past it".

    Everything else — a failed import, a missing optional dependency, a broken venv, an OS
    error — must NOT raise this. Those are the engine being broken, they resolve once it is
    repaired, and they have to stay account-fatal so the cursor is held and the files are
    re-offered. `str(exc)` carries no type information, so without this marker Rust cannot
    tell the two apart and defaults, fail-closed, to "the engine is broken"."""


def _load_model(build):
    """Construct a model, converting ANY failure in OFFLINE mode into ModelNotCached so Rust runs
    the network-allowed `--fetch` helper and retries. Treating every offline construction failure as
    "not downloaded yet" is deliberate: the exception differs by library and version. faster-whisper
    surfaces huggingface_hub's `LocalEntryNotFoundError`, but fastembed 0.8.0 swallows the offline
    error and raises a bare `ValueError("Could not load model ... from any source")` (verified
    against the pinned venv), so matching specific types/messages would silently break first-use
    ingest on a wording change. In offline mode the only recovery for ANY load failure is for the
    fetcher (with the network) to obtain the model and retry, which also self-heals a corrupt cache;
    if it still fails, `request()` fetches once then surfaces the real error. In the FETCHER itself
    (`_OFFLINE` False) this is a passthrough: a genuine download error must surface as-is."""
    if not _OFFLINE:
        return build()
    try:
        return build()
    except ModelNotCached:
        raise
    except Exception as exc:
        raise ModelNotCached(str(exc)) from exc


def get_markitdown():
    global _markitdown
    if _markitdown is None:
        from markitdown import MarkItDown

        _markitdown = MarkItDown()
    return _markitdown


def _pooling_type(name):
    """Map a pooling name from Rust to fastembed's PoolingType (default MEAN)."""
    from fastembed.common.model_description import PoolingType

    return {"mean": PoolingType.MEAN, "cls": PoolingType.CLS}.get(
        (name or "mean").lower(), PoolingType.MEAN
    )


def _local_path_kwargs(spec):
    """`specific_model_path` kwargs for a model that lives on disk, or `{}` for a hub model.

    A locally-trained model (the Stage-4 learned reranker) has no Hugging Face repo to fetch from.
    fastembed's `ModelSource` requires an `hf` or `url` regardless, so Rust registers a deliberate
    placeholder — but `specific_model_path` short-circuits resolution before any source is consulted
    (fastembed 0.8.0, `common/model_management.py`), so the placeholder is never fetched. Loading
    still honours the offline posture: nothing here reaches the network either way.
    """
    path = (spec or {}).get("local_path")
    return {"specific_model_path": path} if path else {}


def get_embedder(model=None, spec=None):
    """The embedder for `model` (default: the English EMBED_MODEL), cached per id.

    `spec` (present for a custom, non-bundled model) carries the fastembed
    `add_custom_model` arguments Rust derived from the registry: the HF source
    repo, the ONNX file, pooling, normalization, and dimension. We register a
    custom model with fastembed once per id (calling it twice raises).
    """
    model = model or EMBED_MODEL
    if model not in _embedders:
        from fastembed import TextEmbedding

        if spec and spec["model"] not in _registered:
            from fastembed.common.model_description import ModelSource

            TextEmbedding.add_custom_model(
                model=spec["model"],
                pooling=_pooling_type(spec.get("pooling")),
                normalization=bool(spec.get("normalize", True)),
                sources=ModelSource(hf=spec["hf"]),
                dim=int(spec["dim"]),
                model_file=spec.get("model_file", "onnx/model.onnx"),
            )
            _registered.add(spec["model"])
        # local_files_only=_OFFLINE makes the offline posture hold even if the env var were somehow
        # not authoritative; fastembed passes it through **kwargs to its huggingface_hub download.
        # cache_dir keeps the weights under the shared model root (issue #286).
        _embedders[model] = _load_model(
            lambda: TextEmbedding(
                model_name=model,
                cache_dir=_fastembed_cache_dir(),
                local_files_only=_OFFLINE,
                **_session_posture(),
                **_local_path_kwargs(spec),
            )
        )
    return _embedders[model]


def get_tokenizer(model=None, spec=None):
    """A tokenizer for sizing chunks by tokens in Rust — an INDEPENDENT COPY of the
    embedder's tokenizer.

    We disable truncation so an oversized chunk reports its *true* length (the splitter
    must see it overflow to break it up). The catch: fastembed's tokenizer is the SAME
    object it embeds with, and it ships with truncation enabled to the model's 512-token
    window. Mutating that shared object with `no_truncation()` (as this used to) stripped
    the embedder's safety net, so a chunk longer than 512 tokens crashed ONNX with a
    `512 vs N` broadcast error instead of being harmlessly truncated. So we CLONE it
    (`from_str(to_str())`) and disable truncation on the copy only — the embedder keeps
    its truncation. The clone reuses the in-memory weights (no download). Padding is left
    as fastembed set it (batch padding on); `do_count_tokens` strips it via the attention
    mask. Falls back to loading the repo tokenizer, then to `None` (caller estimates).
    """
    model = model or EMBED_MODEL
    if model in _tokenizers:
        return _tokenizers[model]
    tok = None
    emb = get_embedder(model, spec)
    candidate = getattr(emb, "tokenizer", None) or getattr(
        getattr(emb, "model", None), "tokenizer", None
    )
    if candidate is not None and hasattr(candidate, "encode"):
        try:
            from tokenizers import Tokenizer

            # Independent copy: disabling truncation here must never touch the embedder's.
            tok = Tokenizer.from_str(candidate.to_str())
            tok.no_truncation()
        except Exception:
            # Clone unavailable: fall through to the repo tokenizer below rather than counting
            # with the SHARED one. That one has truncation ON, so every chunk past the model's
            # window reports exactly the window size — the splitter reads "512, it fits" for a
            # block that does not, and stops splitting the very chunks this exists to catch.
            # A capped count is worse than no count: the chars/4 estimate in `do_count_tokens`
            # is rough, but it keeps growing with the text. Never mutate the shared tokenizer.
            tok = None
    if tok is None:
        try:
            from tokenizers import Tokenizer

            tok = Tokenizer.from_pretrained(model)
            tok.no_truncation()
        except Exception:
            tok = None
    if tok is not None:
        # Cache only on success so a transient failure retries next time.
        _tokenizers[model] = tok
    return tok


def get_reranker(model, spec=None):
    """The cross-encoder reranker for `model`, cached per id.

    `spec` (present for a custom, non-bundled reranker) registers it via
    fastembed's `add_custom_model` from the HF source repo + ONNX file, once per id.
    """
    if model not in _rerankers:
        from fastembed.rerank.cross_encoder import TextCrossEncoder

        if spec and spec["model"] not in _registered:
            from fastembed.common.model_description import ModelSource

            TextCrossEncoder.add_custom_model(
                model=spec["model"],
                sources=ModelSource(hf=spec["hf"]),
                model_file=spec.get("model_file", "onnx/model.onnx"),
            )
            _registered.add(spec["model"])
        _rerankers[model] = _load_model(
            lambda: TextCrossEncoder(
                model_name=model,
                cache_dir=_fastembed_cache_dir(),
                local_files_only=_OFFLINE,
                **_session_posture(),
                **_local_path_kwargs(spec),
            )
        )
    return _rerankers[model]


def get_whisper(model_dir):
    global _whisper
    if _whisper is None:
        from faster_whisper import WhisperModel

        # `model_dir` (passed by Rust) keeps the weights inside PM's data dir so
        # they uninstall cleanly with it; None falls back to the default cache.
        # `local_files_only` mirrors the worker's offline posture: the OFFLINE worker
        # loads from cache only (a miss raises → ModelNotCached), while the `--fetch`
        # helper runs with it False so it can download (issue #286).
        _whisper = _load_model(
            lambda: WhisperModel(
                WHISPER_MODEL,
                device="cpu",
                compute_type="int8",
                download_root=model_dir or None,
                local_files_only=_OFFLINE,
            )
        )
    return _whisper


def clean_text(value):
    """Coerce to tokenizer- and JSON-safe text.

    Converting arbitrary files (and OCR) can yield text carrying lone UTF-16
    surrogates — e.g. when a source file is decoded with surrogate-escape. Those
    are not valid Unicode scalars, so two things downstream reject them: the Rust
    side's strict JSON parser refuses the surrogate escape in our reply (silently
    desyncing the protocol), and HuggingFace `tokenizers` raises "TextEncodeInput
    must be ..." which fails the whole embed batch. Round-tripping through UTF-8
    with errors="ignore" drops only the lone surrogates and leaves all normal
    text untouched. We also coerce non-strings defensively so one odd value can
    never abort a batch.
    """
    if not isinstance(value, str):
        value = "" if value is None else str(value)
    return value.encode("utf-8", "ignore").decode("utf-8", "ignore")


# The largest source file the sidecar will read for conversion / spreadsheet / image analysis
# (F-57). A file past this would balloon the child's memory (the reader materialises it) before the
# 64 MiB response-line cap could ever trip — an OOM on the 8 GB target. Rust pre-flights the same
# limit before it even asks; this is the in-process backstop. Kept in sync with sidecar.rs
# MAX_SIDECAR_INPUT_BYTES.
MAX_INPUT_FILE_BYTES = 128 * 1024 * 1024

# The cap for TEXT-FAMILY files, whose converted Markdown is roughly the size of the input (a .txt
# converts to about itself; a .pdf or .docx extracts to a fraction). Those could clear the 128 MiB
# input cap and then produce a reply that CANNOT fit under the Rust reader's 64 MiB line cap - a
# guaranteed failure, arriving only after minutes of conversion. 40 MiB leaves headroom, since the
# reply is Markdown-wrapped and JSON-escaped and so runs somewhat larger than the source. Kept in
# sync with sidecar.rs MAX_SIDECAR_TEXT_INPUT_BYTES / TEXT_FAMILY_EXTS.
MAX_TEXT_INPUT_FILE_BYTES = 40 * 1024 * 1024
TEXT_FAMILY_EXTS = (".txt", ".md", ".markdown", ".html", ".htm", ".json", ".xml")

# Every modern Office/e-book format is a zip, and its reader inflates the XML inside. The 128 MiB
# input cap above bounds the on-disk (compressed) size, NOT the inflated size, so a zip bomb could
# still balloon memory past it. Reject an archive whose declared uncompressed size, or inflation
# ratio, is implausible for a real document before any reader opens it (H-1 subset).
#
# This started as an .xlsx-only guard, which left .docx / .pptx / .epub — all zips, all handed
# straight to MarkItDown — with nothing but the on-disk cap. The numbers are unchanged: 1 GiB is a
# generous ceiling for any real document, and a genuine Office file inflates ~5-20x (media inside
# is already compressed), nowhere near 200x.
MAX_INFLATED_ARCHIVE_BYTES = 1024 * 1024 * 1024  # 1 GiB total uncompressed
MAX_ARCHIVE_INFLATION_RATIO = 200  # uncompressed : compressed

# The zip-container extensions this guard applies to. .doc/.ppt/.rtf are not zips (OLE / text) and
# .pdf has its own object streams, which MarkItDown's pdfminer bounds itself.
ZIP_CONTAINER_EXTS = (".docx", ".pptx", ".epub", ".xlsx", ".xlsm")


def _guard_file_size(path):
    """Refuse an over-cap input file (F-57) with a clean error instead of OOMing the child. main()
    turns the raised ValueError into an `{ok: false, error}` reply the Rust side surfaces. A missing
    or unreadable file is NOT this guard's concern — it returns quietly so the handler's own path
    handling (e.g. analyze_image's graceful nulls) still applies."""
    import os

    try:
        size = os.path.getsize(path)
    except OSError:
        return
    cap = (
        MAX_TEXT_INPUT_FILE_BYTES
        if str(path).lower().endswith(TEXT_FAMILY_EXTS)
        else MAX_INPUT_FILE_BYTES
    )
    if size > cap:
        raise ValueError(
            f"file is too large to process ({size // (1024 * 1024)} MiB; "
            f"the limit is {cap // (1024 * 1024)} MiB)"
        )


# --- What a document says about ITSELF ---------------------------------------------------------
#
# A filesystem knows no author (#701) — but the FILE very often does. An OOXML container carries
# `docProps/core.xml` (`dc:creator`, `cp:lastModifiedBy`, `dcterms:created`) and a PDF carries an
# Info dictionary (`/Author`, `/CreationDate`). Reading those is what lets a local document show the
# same facts as a Drive one, without PM ever naming the OS account looking at the screen (#709).
#
# Deliberately NOT read: the document's own MODIFIED date. PM's `source_modified_at` for a local
# file is the filesystem mtime, which is exactly what the connector diffs to notice a change;
# handing back a second, differently-sourced "modified" invites precisely that substitution, and a
# document whose docProps date predates the copy on disk would then look permanently stale.
#
# Every reader below returns nulls instead of raising. A property is a nicety — it must never be the
# reason a file fails to land — so the shape is fixed and the failure is silent by construction.

# The largest `docProps/core.xml` this reader will inflate. A real one is a couple of KiB; anything
# past this is a bomb aimed at the single member we open, which `_guard_archive_inflation`'s
# whole-archive 1 GiB ceiling is far too coarse to catch.
MAX_DOC_PROPS_BYTES = 1024 * 1024

# One property value's cap. An author is a name; a hostile file could put the whole member in there,
# and a table cell holding 200 KiB of text helps nobody.
MAX_PROPERTY_CHARS = 256

# The zip containers with a `docProps/core.xml`. .epub is a zip too but stores its metadata in an
# OPF whose path is itself indirected through META-INF/container.xml — a different reader, not
# worth one here until an .epub actually needs it.
OOXML_PROPERTY_EXTS = (".docx", ".pptx", ".xlsx", ".xlsm")

CORE_PROPS_PATH = "docProps/core.xml"
_DC_NS = "{http://purl.org/dc/elements/1.1/}"
_CP_NS = "{http://schemas.openxmlformats.org/package/2006/metadata/core-properties}"
_DCTERMS_NS = "{http://purl.org/dc/terms/}"


def no_properties():
    """The reply when nothing could be read: every key present, every value null. Fixed-shape on
    purpose — the Rust side never has to tell "the reader gave up" from "the document didn't say",
    because for its purposes those are the same answer, and both render "Unknown"."""
    return {"author": None, "last_modified_by": None, "created": None}


def clean_property(value):
    """One property value, or None. Blank and whitespace-only read as "not stated": Word writes
    `<dc:creator/>` for an unset author, and an empty string rendered as an author would be worse
    than "Unknown"."""
    if not isinstance(value, str):
        return None
    text = clean_text(value).strip()
    return text[:MAX_PROPERTY_CHARS] if text else None


def iso_or_none(text):
    """A W3CDTF timestamp as OOXML writes it (`2026-01-04T12:00:00Z`), or None.

    Validated rather than trusted: this string lands in a database column that three date surfaces
    parse, and a malformed one renders as "Invalid Date" on all of them."""
    if not text:
        return None
    import datetime

    try:
        datetime.datetime.fromisoformat(text.replace("Z", "+00:00"))
    except ValueError:
        return None
    return text


def pdf_string(value):
    """A PDF string object as text, or None.

    Values arrive as bytes in either UTF-16 (BOM-prefixed, which is how any non-ASCII name is
    written) or PDFDocEncoding, which agrees with latin-1 across everything a name realistically
    contains. Pure, so it is tested without a PDF library present."""
    if isinstance(value, str):
        return value
    if not isinstance(value, bytes):
        return None
    if value[:2] in (b"\xfe\xff", b"\xff\xfe"):
        return value.decode("utf-16", "replace")
    return value.decode("latin-1", "replace")


def pdf_offset(tz):
    """The UTC offset of a PDF date (`+01'00'`, `Z`, or absent) as an ISO suffix, or None when it is
    not one at all.

    Absent means the spec's "local time, locale unknown". Stored as `Z` so PM keeps ONE timestamp
    format everywhere; the worst case is a displayed DATE shifting by a few hours at a boundary,
    against a stored naive time that no reader could place either."""
    if tz in ("", "Z", "z"):
        return "Z"
    if tz[0] not in "+-":
        return None
    rest = tz[1:].replace("'", "")
    if len(rest) == 2 and rest.isdigit():
        return f"{tz[0]}{rest}:00"
    if len(rest) == 4 and rest.isdigit():
        return f"{tz[0]}{rest[:2]}:{rest[2:]}"
    return None


def pdf_date_to_iso(text):
    """A PDF date string (`D:20260104120000+01'00'`) as ISO-8601, or None.

    Every field after the year is optional in the spec and real writers omit them, so a bare
    `D:2026` is legal and means the first instant of that year. Anything that is not that shape is
    refused outright rather than salvaged — an ISO date wrongly stored in the Info dictionary would
    otherwise be truncated to its year and reported with false precision. Pure."""
    if not text:
        return None
    import datetime

    s = text.strip()
    if s.startswith("D:"):
        s = s[2:]
    lead = 0
    while lead < len(s) and s[lead].isdigit():
        lead += 1
    # 4/6/8/10/12/14 = the spec's truncation points (year … second). Any other run is not a date.
    if lead not in (4, 6, 8, 10, 12, 14):
        return None
    digits, offset = s[:lead], pdf_offset(s[lead:])
    if offset is None:
        return None

    def part(start, end, default):
        piece = digits[start:end]
        return int(piece) if len(piece) == end - start else default

    try:
        stamp = datetime.datetime(
            int(digits[:4]),
            part(4, 6, 1),
            part(6, 8, 1),
            part(8, 10, 0),
            part(10, 12, 0),
            part(12, 14, 0),
        )
    except ValueError:
        return None
    return stamp.strftime("%Y-%m-%dT%H:%M:%S") + offset


def _ooxml_properties(path):
    """`docProps/core.xml` out of an OOXML container, without inflating anything else."""
    try:
        from defusedxml.ElementTree import fromstring
    except Exception:
        # Fail CLOSED, unlike the best-effort `defuse_stdlib()` at the top of this file: that one
        # hardens a parse that ingest depends on, so proceeding beats breaking. This parse is
        # optional, so a venv missing defusedxml simply doesn't get properties.
        return no_properties()
    import zipfile

    try:
        _guard_archive_inflation(path)
        with zipfile.ZipFile(path) as zf:
            info = zf.getinfo(CORE_PROPS_PATH)
            if info.file_size > MAX_DOC_PROPS_BYTES:
                return no_properties()
            raw = zf.read(CORE_PROPS_PATH)
        root = fromstring(raw)
    except Exception:
        # KeyError (a writer that emitted no core.xml — legal), BadZipFile, the inflation guard's
        # ValueError, a parse error, an unreadable file. All the same answer here.
        return no_properties()

    def text(tag):
        el = root.find(tag)
        return clean_property(el.text if el is not None else None)

    return {
        "author": text(f"{_DC_NS}creator"),
        "last_modified_by": text(f"{_CP_NS}lastModifiedBy"),
        "created": iso_or_none(text(f"{_DCTERMS_NS}created")),
    }


def _pdf_properties(path):
    """The Info dictionary of a PDF. Parses the trailer and one object — not the page content — so
    this stays cheap next to the conversion that follows."""
    try:
        from pdfminer.pdfdocument import PDFDocument
        from pdfminer.pdfparser import PDFParser
    except Exception:
        return no_properties()
    try:
        with open(path, "rb") as fh:
            doc = PDFDocument(PDFParser(fh))
            info = doc.info[0] if doc.info else {}
    except Exception:
        # Encrypted, truncated, or not a PDF at all. The conversion that follows delivers the real
        # verdict on the file; this reader has no opinion to add.
        return no_properties()
    return {
        "author": clean_property(pdf_string(info.get("Author"))),
        # A PDF names a Producer and a Creator, both of which are APPLICATIONS rather than people.
        # Rendering "Microsoft Word" under "Modified by" would be a wrong answer, which is strictly
        # worse than the missing one.
        "last_modified_by": None,
        "created": pdf_date_to_iso(pdf_string(info.get("CreationDate"))),
    }


def read_document_properties(path):
    """Author, last editor and creation date as the DOCUMENT states them, for the formats that
    state them. Always `{author, last_modified_by, created}`; each value null when unstated."""
    lower = str(path).lower()
    if lower.endswith(OOXML_PROPERTY_EXTS):
        return _ooxml_properties(path)
    if lower.endswith(".pdf"):
        return _pdf_properties(path)
    # .txt/.md/.html/.csv and the rest carry no property block at all. Not a failure: most files
    # genuinely have no author to state, which is the case "Unknown" exists for.
    return no_properties()


def do_file_properties(params):
    """What one file says about itself. Rust asks only for the extensions that can answer, so an
    all-null reply here means the document really didn't say."""
    path = params["path"]
    try:
        _guard_file_size(path)
    except ValueError:
        # Over the input cap: `convert` refuses this file anyway, so there is no document for these
        # properties to belong to.
        return no_properties()
    return read_document_properties(path)


def convert_local_tolerating_charset(engine, path, stream_info_cls):
    """`engine.convert_local(path)`, retried once as UTF-8 if the first pass could not decode.

    MarkItDown guesses a file's charset from its **first 4 KB** (`_get_stream_info_guesses` reads
    `file_stream.read(4096)` and hands it to `charset_normalizer`), then gives that one guess to
    every converter, each of which decodes the WHOLE file with it. So a file whose opening 4 KB
    happens to be plain ASCII is labelled `ascii`, and a single accented character further in
    raises `UnicodeDecodeError`.

    That would merely lose one converter, except the raise comes out of `accepts()` — the sniffing
    pass — and MarkItDown's `_convert` guards `accepts()` against `NotImplementedError` and nothing
    else. One converter declining to sniff therefore takes down the converters that would have
    succeeded: a 27 KB JSON file with an em dash in it is refused by the *notebook* converter's
    sniff, and never reaches the plain-text one.

    Retrying as UTF-8 is safe rather than optimistic: **ASCII is a strict subset of UTF-8**, so
    re-reading can only widen what decodes — it cannot corrupt a file the first pass would have
    read, and it cannot rescue one that is genuinely some other encoding. Those fail again and are
    skipped by the caller, which is the same outcome as before, minus the pinned cursor.

    `stream_info_cls` is passed in rather than imported so this stays testable without markitdown
    installed — the gate's Python has no venv (see `get_markitdown`'s lazy import).
    """
    try:
        return engine.convert_local(path)
    except UnicodeDecodeError:
        return engine.convert_local(path, stream_info=stream_info_cls(charset="utf-8"))


def do_convert(params):
    """Convert one file to Markdown. Returns its text and a best-effort title."""
    path = params["path"]
    # PM's own caps are verdicts about THIS file, so they carry the `unconvertible` marker.
    try:
        _guard_file_size(path)
        if str(path).lower().endswith(ZIP_CONTAINER_EXTS):
            _guard_archive_inflation(path)
    except ValueError as exc:
        raise Unconvertible(str(exc)) from exc

    # Deliberately OUTSIDE the try below: `get_markitdown()` imports markitdown lazily, so a
    # broken or half-installed venv surfaces HERE as an ImportError/ModuleNotFoundError. That is
    # the engine being broken, not a verdict on this file — it must stay account-fatal so the
    # delta cursor is held and every affected file is re-offered once the engine is repaired.
    engine = get_markitdown()
    # Imported HERE, not at module scope: markitdown is a deliberately lazy import (see
    # `get_markitdown`), and naming its exceptions at module level would load the whole package
    # on every sidecar start — including for `embed`, `rerank` and `transcribe`, which never
    # convert anything. Safe at this point: the engine has already imported successfully.
    from markitdown import MarkItDownException, MissingDependencyException, StreamInfo

    # `convert_local`, NOT `convert`: MarkItDown's `convert` dispatches on the STRING, and a path
    # beginning `http:` / `https:` / `file:` / `data:` is routed to `convert_uri` — a network fetch
    # from the one process that must never hold a socket. The names we pass are staged copies or
    # the user's own paths, so this is not reachable today; the local entry point makes it
    # unreachable by construction rather than by luck.
    try:
        result = convert_local_tolerating_charset(engine, path, StreamInfo)
    except MissingDependencyException:
        # A MarkItDownException, but NOT a verdict on the file: this format needs an optional
        # package the venv is missing. Repairing the engine fixes it, so let it stay account-fatal
        # rather than skipping every file of that type forever.
        raise
    except (MarkItDownException, UnicodeDecodeError) as exc:
        # UnsupportedFormatException / FileConversionException: the engine read this file and
        # refused it. Retrying forever is what pins the account, so this one is a skip.
        #
        # A `UnicodeDecodeError` arriving here has already survived the UTF-8 retry above, so the
        # file really is in some encoding the engine cannot read — a verdict, not a broken engine.
        # Before it was listed here it escaped to `main`'s catch-all and came back WITHOUT
        # `error_kind`, which Rust reads as an engine fault: the item was retried forever and the
        # delta cursor never moved past it.
        raise Unconvertible(str(exc)) from exc
    title = (getattr(result, "title", None) or "").strip()
    return {
        "markdown": clean_text(result.text_content or ""),
        "title": clean_text(title),
    }


def do_embed(params):
    """Embed a batch of strings with the selected ONNX model.

    Rust prepends any asymmetric retrieval prefix (e5's `query:` / `passage:`)
    before sending, so the text is embedded as-is. `model`/`custom` select and
    (for a custom model) register the embedder; both default to the English model.
    `batch_size`, when present, caps how many texts the embedder processes per
    forward pass — the "gentle" indexing lever, which bounds peak activation
    memory on a low-memory machine. Absent → fastembed's own default batch.
    """
    # Sanitize before tokenizing: one chunk with a stray surrogate would
    # otherwise fail the entire batch (and thus the whole document).
    texts = [clean_text(t) for t in params.get("texts", [])]
    model = params.get("model") or EMBED_MODEL
    spec = params.get("custom")
    embedder = get_embedder(model, spec)
    batch_size = params.get("batch_size")
    # fastembed yields numpy arrays; hand back plain lists so they serialize.
    if batch_size:
        vecs = embedder.embed(texts, batch_size=int(batch_size))
    else:
        vecs = embedder.embed(texts)
    vectors = [vec.tolist() for vec in vecs]
    dim = int(spec["dim"]) if spec and spec.get("dim") else EMBED_DIM
    return {"vectors": vectors, "dim": dim, "model": model}


def do_count_tokens(params):
    """Token counts for a batch, using the selected embedder's own tokenizer.

    The Rust splitter sizes chunks by tokens (never chars) so a chunk never
    overflows the embedder's input window. Counting with the same tokenizer that
    embeds keeps the two in lockstep. If no tokenizer can be loaded, fall back to a
    rough chars/4 estimate so chunking still works rather than failing the document.
    """
    texts = [clean_text(t) for t in params.get("texts", [])]
    model = params.get("model") or EMBED_MODEL
    spec = params.get("custom")
    tok = get_tokenizer(model, spec)
    if tok is not None:
        try:
            encodings = tok.encode_batch(texts)
            # `encode_batch` pads every text to the batch's longest (fastembed
            # leaves batch padding enabled, and this is the same shared tokenizer
            # it embeds with, so we must not turn padding off here). `len(e.ids)`
            # would then count pad tokens and report every text as the longest
            # one's length — which made the splitter size every block by the
            # document's largest block and shatter long documents. The attention
            # mask is 1 for real tokens and 0 for padding, so summing it gives each
            # text's TRUE length, independent of the rest of the batch.
            return {"counts": [int(sum(e.attention_mask)) for e in encodings]}
        except Exception:
            pass
    return {"counts": [max(1, len(t) // 4) for t in texts]}


def do_rerank(params):
    """Score query<->passage relevance with the selected cross-encoder reranker.

    Returns one score per passage (higher = more relevant); Rust reorders by them.
    `model`/`custom` select and (for a custom model) register the reranker. Passage
    text is untrusted data — scored, never executed.
    """
    query = clean_text(params.get("query", ""))
    passages = [clean_text(p) for p in params.get("passages", [])]
    if not passages:
        return {"scores": []}
    reranker = get_reranker(params.get("model"), params.get("custom"))
    scores = list(reranker.rerank(query, passages))
    return {"scores": [float(s) for s in scores]}


def do_transcribe(params):
    """Transcribe one audio clip to text with the local Whisper model.

    The Rust side records the clip in the webview, writes the bytes to a temp
    file, and passes its path here; faster-whisper decodes the audio itself via
    the bundled PyAV (no system ffmpeg needed). The text is the user's own speech
    bound for the chat box — still untrusted data, so it is `clean_text`-ed like
    every other string we return.
    """
    path = params["path"]
    model = get_whisper(params.get("model_dir"))
    segments, _info = model.transcribe(path)
    text = " ".join(segment.text.strip() for segment in segments).strip()
    return {"text": clean_text(text)}


def _forbid_rapidocr_downloads():
    """Make the OFFLINE posture authoritative for rapidocr, which has no offline flag of its own.

    Every other model reaches the network through huggingface_hub, so `HF_HUB_OFFLINE` (set at
    import) gates it. rapidocr does not: it calls its OWN downloader against its OWN host, which
    no HF variable touches. Left alone, a cold cache makes the worker that is parsing untrusted
    file bytes open an outbound socket mid-ingest — the exact egress issue #286 forbids, and the
    one the sandbox's fall-open contract cannot be relied on to stop.

    So we replace the downloader's single entry point with one that refuses to fetch. A cold cache
    then surfaces as a clean `ModelNotCached`, `request()` runs the network-allowed `--fetch`
    helper, and the retry finds the models on disk — which is how the embedder, reranker and speech
    model already work.

    It has to keep the "already downloaded" case working, and that is not free: rapidocr's engine
    calls the downloader UNCONDITIONALLY and lets the downloader decide whether the file is already
    there. Refusing outright would therefore reject a warm cache too — the retry would fail exactly
    like the first attempt and OCR would stay off for good. So the replacement keeps that one
    branch: the file is there, return; it is not, report a miss. Nothing else about it can reach
    the network, and no checksum is re-verified here — a corrupt model fails the ONNX load instead,
    which `_load_model` turns into the same miss, and the fetcher then re-downloads it against the
    checksum it holds.

    Best-effort by design: if rapidocr's internals move, OCR must not break outright. The
    constructor would then attempt its own download, fail under confinement, and `_load_model`
    still reports the miss — the same recovery, one round trip later.
    """
    try:
        from rapidocr.utils.download_file import DownloadFile

        def _refuse(params):
            save_path = getattr(params, "save_path", None)
            if save_path is not None and os.path.exists(save_path):
                return
            raise ModelNotCached("rapidocr's OCR models are not downloaded yet")

        DownloadFile.run = staticmethod(_refuse)
    except Exception:
        pass


def get_ocr_engine():
    """The OPTIONAL on-device OCR engine (rapidocr), cached. Lazy-imported so the engine — and its
    image-processing deps (opencv/shapely/pyclipper) — only load when OCR is actually requested, and
    so the base sidecar runs without them installed. `RapidOCR` runs on the SAME onnxruntime that
    fastembed already ships. Raises ImportError if the optional component isn't installed — the
    Rust side only requests OCR once the install is confirmed, so that never fires in normal use.

    Its models live under `PM_MODELS_DIR` beside every other model, and the construction goes
    through `_load_model` so a cold cache in the OFFLINE worker becomes `ModelNotCached` and rides
    the fetch-and-retry path. Before that, this was the one loader outside that contract: under the
    no-network confinement rapidocr could never obtain its models, the constructor raised on every
    photo, `do_analyze_image` swallowed it, and OCR was permanently and silently off.
    """
    global _ocr_engine
    if _ocr_engine is None:
        from rapidocr import RapidOCR

        if _OFFLINE:
            _forbid_rapidocr_downloads()
        model_dir = _rapidocr_model_dir()
        params = {"Global.model_root_dir": model_dir} if model_dir else None
        _ocr_engine = _load_model(lambda: RapidOCR(params=params))
    return _ocr_engine


def _open_image(path):
    """Open an image with Pillow, registering the HEIC opener if pi-heif is available. Pillow is
    already present (a markitdown dependency); pi-heif ships in the optional OCR component, so
    HEIC decoding needs that component installed — a non-HEIC image opens regardless.

    pi-heif rather than pillow-heif: same bindings, same author, same repository, but the
    decode-only build. See OPTIONAL_OCR_PINS in sidecar.rs for why the encoder build had to go.
    A venv still holding the old pillow-heif is not read here — changing the pins changes the
    component's stamp, so PM reports photo OCR as not installed until it is re-added.
    """
    global _heif_registered
    from PIL import Image

    if not _heif_registered:
        try:
            from pi_heif import register_heif_opener

            register_heif_opener()
        except Exception:
            pass  # pi-heif absent → HEIC won't open, every other format still does
        _heif_registered = True
    return Image.open(path)


def _gps_to_decimal(value, ref):
    """Convert an EXIF GPS coordinate (a (deg, min, sec) triple of rationals) + its hemisphere ref
    ('N'/'S'/'E'/'W') to a signed decimal degree, or None if absent/malformed.
    """
    if not value or ref is None:
        return None
    try:
        deg, minute, sec = (float(v) for v in value)
    except Exception:
        return None
    dec = deg + minute / 60.0 + sec / 3600.0
    if str(ref).strip().upper() in ("S", "W"):
        dec = -dec
    # A GPS rational with a zero denominator makes float() produce nan/inf, and a non-finite float
    # serializes as bare `NaN`/`Infinity` -- which is not valid JSON, so the Rust reader skips the
    # reply line and blocks on an answer that never comes. One corrupt-EXIF photo would wedge the
    # whole serialized sidecar (ingest, retrieval, rerank, transcribe, map) until the per-method
    # timeout expired. A photo with no readable location is the honest answer here.
    if not math.isfinite(dec):
        return None
    return round(dec, 6)


def _exif_meta(img):
    """Best-effort (capture_date, lat, lon) from an image's EXIF. capture_date is the EXIF
    DateTimeOriginal normalised to 'YYYY-MM-DD' (the GPS/Exif sub-IFDs are where phones keep these);
    any field absent → None, and the Rust side fills capture_date from the filename/ingest time.
    """
    capture_date = lat = lon = None
    try:
        exif = img.getexif()
    except Exception:
        return capture_date, lat, lon
    if not exif:
        return capture_date, lat, lon

    # DateTimeOriginal (0x9003) is in the Exif sub-IFD (0x8769); fall back to base DateTime (306).
    dt = None
    try:
        dt = exif.get_ifd(0x8769).get(0x9003)
    except Exception:
        pass
    dt = dt or exif.get(306)
    # EXIF stores "YYYY:MM:DD HH:MM:SS"; take the date and switch ':' to '-'. Guard against the
    # all-zero placeholder some cameras write.
    if isinstance(dt, str) and len(dt) >= 10 and dt[:4].isdigit() and dt[:4] != "0000":
        capture_date = dt[:10].replace(":", "-")

    try:
        gps = exif.get_ifd(0x8825)  # GPS IFD: 1=LatRef 2=Lat 3=LonRef 4=Lon
    except Exception:
        gps = None
    if gps:
        lat = _gps_to_decimal(gps.get(2), gps.get(1))
        lon = _gps_to_decimal(gps.get(4), gps.get(3))
    return capture_date, lat, lon


def _run_ocr(engine, img, path):
    """Recognise text in an image, returning the joined lines (newest rapidocr returns an object
    with a `.txts` tuple; tolerate the older list-of-[box,text,score] shape too). Pass a decoded RGB
    array when we already opened the image (HEIC works); else hand rapidocr the path directly.

    rapidocr prints init/inference chatter, which would corrupt the reply channel. `main()` now
    redirects stdout around EVERY handler, so this call — and the engine construction — no longer
    carry a guard of their own.
    """
    if img is not None:
        import numpy as np

        target = np.array(img.convert("RGB"))
    else:
        target = path  # let rapidocr read the file itself (non-HEIC)
    result = engine(target)

    txts = getattr(result, "txts", None)
    if txts:
        return "\n".join(t for t in txts if t)
    # Older shape: ([[box, text, score], ...], elapse) or [[box, text, score], ...].
    rows = result
    if isinstance(result, tuple) and result and isinstance(result[0], (list, tuple)):
        rows = result[0]
    if isinstance(rows, (list, tuple)):
        out = [str(r[1]) for r in rows if isinstance(r, (list, tuple)) and len(r) >= 2]
        return "\n".join(out)
    return ""


def do_analyze_image(params):
    """Analyse one image for photo ingestion: EXIF capture metadata (always) + OCR text (only when
    `run_ocr`). The Rust side sets `run_ocr` from whether the optional component is installed, so
    a user who declined OCR still gets dimensions + EXIF (date/location) for the metadata chunk.
    OCR output is untrusted text — `clean_text`-ed like every other string we return. An unreadable
    image (e.g. HEIC without pi-heif) yields nulls; Rust falls back to a filename/ingest date.
    """
    path = params["path"]
    _guard_file_size(path)
    run_ocr = bool(params.get("run_ocr", False))
    width = height = None
    capture_date = lat = lon = None
    img = None
    try:
        img = _open_image(path)
        width, height = img.size
        capture_date, lat, lon = _exif_meta(img)
    except Exception:
        pass  # unreadable / HEIC without codec — metadata stays null, OCR may still try the path

    ocr_text = ""
    ocr_ran = False
    if run_ocr:
        try:
            ocr_text = _run_ocr(get_ocr_engine(), img, path)
            ocr_ran = True
        except ModelNotCached:
            # NOT a broken component: the offline worker just needs the fetcher to download
            # rapidocr's models, after which `request()` retries this very call. Degrading here
            # would commit a photo that claims to hold no text and never look again — the whole
            # point of routing OCR through the fetch-and-retry contract. Must precede the broad
            # `except` below; ModelNotCached is an Exception too.
            raise
        except Exception as exc:
            # F-56: a broken/offline OCR component must degrade to EXIF-only, not fail the whole
            # photo ingest. The metadata gathered above is intact; Rust reads ocr_ran=false.
            sys.stderr.write(f"pm_sidecar: OCR failed for {path!r}, keeping EXIF only: {exc}\n")
            sys.stderr.flush()
            ocr_text = ""
            ocr_ran = False

    return {
        "ocr_text": clean_text(ocr_text),
        "ocr_ran": ocr_ran,
        "capture_date": capture_date,
        "lat": lat,
        "lon": lon,
        "width": width,
        "height": height,
    }


# ---- spreadsheet ingestion (dedicated processor; bypasses MarkItDown) --------
#
# .xlsx/.csv are routed here by Rust instead of through MarkItDown, which would
# flatten them into one Markdown pipe table the generic chunker then slices badly
# (rows lose their header context). We read VALUES ONLY — no formula evaluation, no
# styling — and hand Rust a per-sheet structure it shapes into a metadata chunk plus
# self-describing row chunks. Column types come from a lightweight try-parse heuristic
# (`infer_column_type`), deliberately NOT pandas dtype inference, which is unreliable
# on messy real-world CSVs. Cells are stringified for output so the Rust side receives
# plain text and never sees a Python type.

# Upper bound on rows returned PER SHEET. A sheet with more rows is truncated to this
# many (its `row_count` still reports the TRUE total and `truncated` is set), so a
# 200k-row sheet can't explode into 200k row chunks + embeddings. Tunable; composes
# with — does not replace — the byte-size cap Rust already enforces on a fetched body.
SPREADSHEET_ROW_CAP = 5000

# Upper bound on columns returned PER SHEET. The row cap alone bounds one dimension: a sheet 4,000
# columns wide still fits under it and still produces a reply the Rust reader must reject at its
# 64 MiB line cap — after minutes of parsing, having read the file twice. Headers past this are
# dropped; `col_count` still reports the TRUE total and `cols_truncated` is set. 256 is far past
# any spreadsheet a person reads (Excel's own A-IV classic limit), so a real sheet never meets it.
SPREADSHEET_COL_CAP = 256

# Upper bound on the characters of ONE cell. A single cell can legally hold 32k characters, and a
# pasted-in wall of text tells a row chunk nothing the first line didn't.
SPREADSHEET_CELL_CHARS = 1000

# The character budget for one reply's cell text, SHARED across every sheet in the workbook. The
# row and column caps bound each sheet's shape but not the product (5,000 x 256 x 1,000 characters
# is three orders of magnitude past the reader's line cap), and a workbook can hold any number of
# sheets. This is the one bound that actually holds: rows stop being kept once it is spent, and the
# sheet reports itself `truncated` exactly as a row-capped one does. 8 MiB of cell text leaves
# generous headroom under the 64 MiB line cap once JSON-escaped, and a real 5,000-row sheet of 20
# columns spends well under 1 MiB.
SPREADSHEET_TEXT_BUDGET = 8 * 1024 * 1024

# How many non-empty values per column the type heuristic samples — enough to be
# confident without scanning a whole column.
_TYPE_SAMPLE = 50

_DATE_FORMATS = (
    "%Y-%m-%d",
    "%Y/%m/%d",
    "%d-%m-%Y",
    "%d/%m/%Y",
    "%m/%d/%Y",
    "%Y-%m-%dT%H:%M:%S",
    "%Y-%m-%d %H:%M:%S",
)


def _parse_date(value):
    """A `datetime.date` if `value` is a date/datetime or a string in a common format, else None.
    Shared by column-type inference and the sheet's date range."""
    import datetime

    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    if isinstance(value, str):
        s = value.strip()
        for fmt in _DATE_FORMATS:
            try:
                return datetime.datetime.strptime(s, fmt).date()
            except ValueError:
                continue
    return None


def _cell_type(value):
    """Classify one cell into 'int' | 'float' | 'date' | 'bool' | 'string', or None when empty.
    Native types from openpyxl (int/float/datetime/bool) are honoured directly; a string (all
    a CSV yields) is try-parsed. Empty/None returns None so it never sways the column's type."""
    import datetime

    if value is None:
        return None
    if isinstance(value, bool):
        return "bool"
    if isinstance(value, int):
        return "int"
    if isinstance(value, float):
        return "float"
    if isinstance(value, (datetime.date, datetime.datetime)):
        return "date"
    s = str(value).strip()
    if not s:
        return None
    if s.lower() in ("true", "false", "yes", "no"):
        return "bool"
    try:
        int(s)
        return "int"
    except ValueError:
        pass
    try:
        float(s)
        return "float"
    except ValueError:
        pass
    if _parse_date(s) is not None:
        return "date"
    return "string"


def infer_column_type(values):
    """The dominant inferred type over a column's non-empty values — 'int' | 'float' | 'date' |
    'bool' | 'string' | 'empty'. Standalone and pure: it takes raw cell values and returns a label
    with NO coupling to how chunks are emitted, so a later card can reuse it to detect Status/Date-
    like columns for a different purpose. A numeric column mixing ints and floats reads as 'float';
    a genuinely heterogeneous column reads as 'string' (the safe catch-all)."""
    seen = set()
    for v in values:
        t = _cell_type(v)
        if t is not None:
            seen.add(t)
    if not seen:
        return "empty"
    if len(seen) == 1:
        return next(iter(seen))
    if seen <= {"int", "float"}:
        return "float"
    return "string"


def inspect_columns(headers, rows):
    """A standalone schema descriptor for a sheet: one {name, inferred_type} per column, inferred
    from a sample of each column's values. Decoupled from chunk emission (that lives in Rust), so
    this is the single reusable place column names + types are derived — a later card reuses it."""
    columns = []
    for i, name in enumerate(headers):
        sample = []
        for row in rows:
            if i < len(row) and row[i] is not None and str(row[i]).strip():
                sample.append(row[i])
                if len(sample) >= _TYPE_SAMPLE:
                    break
        columns.append({"name": clean_text(name), "inferred_type": infer_column_type(sample)})
    return columns


def _date_range(rows, columns):
    """[min, max] 'YYYY-MM-DD' over the first date-typed column, or None if no column is a date."""
    idx = next((i for i, c in enumerate(columns) if c["inferred_type"] == "date"), None)
    if idx is None:
        return None
    dates = []
    for row in rows:
        if idx < len(row):
            d = _parse_date(row[idx])
            if d is not None:
                dates.append(d)
    if not dates:
        return None
    return [min(dates).isoformat(), max(dates).isoformat()]


class _TextBudget:
    """The remaining cell-text characters for one reply, shared across a workbook's sheets.

    Deliberately a tiny mutable object rather than a running total threaded through return values:
    the budget is a property of the REPLY, and every sheet spends from the same pot. `spend`
    returns False once a row would not fit, and stays False afterwards — a later, smaller row must
    not sneak in and leave `rows` non-contiguous with the sheet it came from.
    """

    def __init__(self, total):
        self.remaining = total

    def spend(self, chars):
        if chars > self.remaining:
            self.remaining = 0
            return False
        self.remaining -= chars
        return True


def _cell_text(value):
    """One cell as the string Rust receives: cleaned and capped (SPREADSHEET_CELL_CHARS)."""
    if value is None:
        return ""
    return clean_text(str(value))[:SPREADSHEET_CELL_CHARS]


def _build_sheet(name, rows_iter, budget):
    """Consume one sheet's rows (lazily — the caller keeps the file/workbook open): the first row is
    the header, the rest are data. Keeps up to SPREADSHEET_ROW_CAP data rows, SPREADSHEET_COL_CAP
    columns, and whatever `budget` has left, while counting the TRUE row and column totals — so a
    huge sheet reports honestly and is flagged `truncated` / `cols_truncated`. Returns the
    per-sheet dict Rust consumes, or None for an empty sheet (no header row)."""
    it = iter(rows_iter)
    try:
        header_row = next(it)
    except StopIteration:
        return None
    col_count = len(header_row)
    headers = [_cell_text(c) for c in list(header_row)[:SPREADSHEET_COL_CAP]]

    kept = []  # native cell rows (capped) — inferred on before stringifying, for accurate types
    str_rows = []  # the same rows as the strings actually sent, which is what the budget measures
    total = 0
    spent_out = False
    for row in it:
        total += 1
        if spent_out or len(kept) >= SPREADSHEET_ROW_CAP:
            # Keep counting: `row_count` is the sheet's TRUE total, which is what makes the
            # truncation note honest rather than a silently short table.
            continue
        native = list(row)[:SPREADSHEET_COL_CAP]
        cells = [_cell_text(c) for c in native]
        if not budget.spend(sum(len(c) for c in cells)):
            spent_out = True
            continue
        kept.append(native)
        str_rows.append(cells)

    columns = inspect_columns(headers, kept)
    date_range = _date_range(kept, columns)
    return {
        "name": clean_text(name),
        "headers": headers,
        "col_count": col_count,
        "cols_truncated": col_count > len(headers),
        "row_count": total,
        "inferred_types": [c["inferred_type"] for c in columns],
        "date_range": date_range,
        "rows": str_rows,
        "truncated": total > len(kept),
    }


def _csv_encoding(path):
    """Pick a text encoding for a CSV: UTF-8 (BOM-aware) if the head decodes, else cp1252 — Excel's
    Windows default (F-55). Prevents a non-UTF-8 export from failing the whole ingest with a
    UnicodeDecodeError. The caller opens with errors='replace' so a stray byte past the sampled head
    still degrades to a replacement char rather than raising mid-parse."""
    with open(path, "rb") as fh:
        head = fh.read(65536)
    try:
        head.decode("utf-8-sig")
        return "utf-8-sig"
    except UnicodeDecodeError:
        return "cp1252"


def _sheets_from_csv(path, budget):
    """A CSV is a single sheet named after the file; delimiter is sniffed (falls back to comma).
    Encoding is UTF-8 with a cp1252 fallback (F-55) so an Excel export doesn't crash ingest."""
    import csv
    import os

    fh = open(path, "r", encoding=_csv_encoding(path), errors="replace", newline="")
    try:
        sample = fh.read(65536)
        fh.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
        name = os.path.splitext(os.path.basename(path))[0]
        sheet = _build_sheet(name, csv.reader(fh, dialect), budget)
        return [sheet] if sheet else []
    finally:
        fh.close()


def _guard_archive_inflation(path):
    """Refuse a zip bomb before any reader inflates it (H-1 subset). Sums the archive's DECLARED
    uncompressed and compressed sizes (cheap — no decompression) and rejects when the total
    uncompressed size, or the inflation ratio, is implausible for a real document. openpyxl's
    `read_only=True` streams rows lazily, so for a workbook the main eager cost this bounds is the
    shared-strings table; MarkItDown's .docx/.pptx/.epub readers materialise their XML outright.
    A non-zip file returns quietly — the reader raises its own clean error."""
    import zipfile

    try:
        with zipfile.ZipFile(path) as zf:
            uncompressed = sum(i.file_size for i in zf.infolist())
            compressed = sum(i.compress_size for i in zf.infolist())
    except (zipfile.BadZipFile, OSError):
        return
    if uncompressed > MAX_INFLATED_ARCHIVE_BYTES:
        raise ValueError(
            f"this file expands to too much data "
            f"({uncompressed // (1024 * 1024)} MiB uncompressed; "
            f"the limit is {MAX_INFLATED_ARCHIVE_BYTES // (1024 * 1024)} MiB)"
        )
    if compressed > 0 and uncompressed // compressed > MAX_ARCHIVE_INFLATION_RATIO:
        raise ValueError(
            f"this file's compression ratio is implausibly high "
            f"({uncompressed // compressed}x; the limit is {MAX_ARCHIVE_INFLATION_RATIO}x)"
        )


def _sheets_from_xlsx(path, budget):
    """Every worksheet in an .xlsx/.xlsm. `data_only=True` returns cached VALUES (never formulas);
    `read_only=True` streams rows so a large workbook stays memory-bounded."""
    from openpyxl import load_workbook

    _guard_archive_inflation(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        out = []
        for ws in wb.worksheets:
            sheet = _build_sheet(ws.title, ws.iter_rows(values_only=True), budget)
            if sheet:
                out.append(sheet)
        return out
    finally:
        wb.close()


def do_analyze_spreadsheet(params):
    """Parse a spreadsheet (.xlsx/.csv) into per-sheet structure for the dedicated ingest path,
    bypassing MarkItDown. Reads VALUES ONLY — no formula evaluation, no styling. Each sheet reports
    its headers, TRUE row_count and col_count, per-column inferred types, an optional date range,
    and the rows that fit the row / column / reply-size caps (`truncated` and `cols_truncated` say
    when they bit). Rust shapes these into a metadata chunk + self-describing row chunks. Cell text
    is untrusted data — cleaned, never executed. Legacy .xls is not supported (its xlrd parser
    surface was dropped, H-1)."""
    path = params["path"]
    _guard_file_size(path)
    ext = (params.get("ext") or "").lower().lstrip(".")
    # One budget for the whole reply: a workbook of many sheets must not multiply past the reader's
    # line cap just because each sheet is individually small.
    budget = _TextBudget(SPREADSHEET_TEXT_BUDGET)
    if ext in ("xlsx", "xlsm"):
        sheets = _sheets_from_xlsx(path, budget)
    elif ext == "csv":
        sheets = _sheets_from_csv(path, budget)
    else:
        raise ValueError(f"unsupported spreadsheet extension: {ext!r}")
    return {"sheets": sheets, "row_cap": SPREADSHEET_ROW_CAP}


def _pca(x, k):
    """Project mean-centred rows `x` (n, d) onto their top-`k` principal axes via SVD.

    numpy only (already a dependency via fastembed), so this path needs no extra install. Used both
    as the default 2-D reducer and as the 50-d initialisation for t-SNE.
    """
    import numpy as np

    k = min(k, x.shape[1], x.shape[0])
    # Economy SVD: columns of Vt are the principal directions; U*S is the projection onto them.
    u, s, _vt = np.linalg.svd(x, full_matrices=False)
    return (u[:, :k] * s[:k]).astype(np.float32)


def _fit_unit(coords):
    """Min-max scale each axis into [0,1] so the Rust/TS side never guesses the coord range."""
    import numpy as np

    c = np.asarray(coords, dtype=np.float32)
    mn = c.min(axis=0)
    mx = c.max(axis=0)
    span = np.where((mx - mn) > 1e-9, mx - mn, 1.0)
    return (c - mn) / span


def _reduce(vecs, method):
    """Reduce per-document vectors to 2-D. Returns (coords, method_actually_used).

    `pca` (the bundled default) is numpy-only and instant. `tsne` uses openTSNE — an OPTIONAL
    component the user downloads from Settings; if it isn't installed (or fails) we fall back to
    PCA and report `pca`, so the map is always usable. The marker the Rust side checks decides what
    to *request*; this decides what actually ran.
    """
    import numpy as np

    x = vecs - vecs.mean(axis=0, keepdims=True)  # mean-centre
    if method != "tsne":
        return _pca(x, 2), "pca"
    try:
        from openTSNE import TSNE

        n, d = x.shape
        # Pre-reduce to 50-d with PCA to speed up the neighbour search (a standard t-SNE step); the
        # 2-D start is PCA-initialised, which keeps the layout deterministic given random_state.
        x_in = _pca(x, 50) if d > 50 else x
        ts = TSNE(
            n_components=2,
            perplexity=min(30, max(5, (n - 1) // 3)),
            metric="cosine",
            initialization="pca",
            n_jobs=1,
            random_state=42,
            verbose=False,
        )
        return np.asarray(ts.fit(x_in), dtype=np.float32), "tsne"
    except Exception:
        # Not installed, or a runtime failure — a usable PCA map beats no map.
        return _pca(x, 2), "pca"


def do_reduce(params):
    """Project a batch of per-document vectors to 2-D for the semantic memory map.

    params: {"vectors": [[float]...], "method": "pca"|"tsne"}
    returns: {"coords": [[x,y]...] in [0,1]^2, "method": <actually used>}

    Untrusted-data rule still holds: numbers in, numbers out, never executed.
    """
    raw = params.get("vectors") or []
    n = len(raw)
    if n == 0:
        return {"coords": [], "method": "none"}
    if n <= 3:
        # t-SNE/PCA degenerate for a handful of points; a deterministic spread keeps the map sane.
        # (These guards stay numpy-free so a tiny map needs no scientific stack at all.)
        return {"coords": [[float(i), 0.0] for i in range(n)], "method": "trivial"}
    import numpy as np

    vecs = np.asarray(raw, dtype=np.float32)
    coords, used = _reduce(vecs, (params.get("method") or "pca").lower())
    return {"coords": _fit_unit(coords).tolist(), "method": used}


def do_net_selftest(_params):
    """Dev-only network-block probe (issue #286): report whether the OS refused (1) a direct
    outbound TCP socket and (2) out-of-process DNS resolution.

    The socket probe targets TEST-NET-1 (192.0.2.1, RFC 5737 — reserved, routes nowhere), so an
    unconfined attempt is egress-safe; confined it is refused (Windows AppContainer → WSAEACCES;
    Linux seccomp → EACCES; macOS `(deny default)` → the socket is denied) before a packet leaves.

    The DNS probe is the macOS-specific gate (finding #1): hostname resolution on macOS goes to the
    mDNSResponder daemon over a mach service a `(deny network*)` rule never sees, so blocking direct
    sockets is not enough — we also attempt `getaddrinfo` and report whether it was refused. A
    confined worker can't reach the resolver (no mach-lookup to mDNSResponder on macOS; no socket at
    all on the other arms), so this must fail; if resolution SUCCEEDS the confinement has an
    out-of-process egress hole. Registered ONLY when PM_SIDECAR_DEV=1, so a release worker has no
    reachable path here."""
    import errno as _errno
    import socket

    wsaeacces = 10013  # Windows: a no-network AppContainer refuses the socket with this.
    try:
        with socket.create_connection(("192.0.2.1", 443), timeout=2):
            blocked, detail, err = False, "connected — network is NOT blocked", None
    except OSError as exc:
        blocked = (
            isinstance(exc, PermissionError)
            or exc.errno == _errno.EACCES
            or getattr(exc, "winerror", None) == wsaeacces
        )
        detail = (
            "outbound socket refused (network blocked)"
            if blocked
            else f"reached the network layer, NOT blocked ({type(exc).__name__}: {exc})"
        )
        err = exc.errno

    try:
        socket.getaddrinfo("example.com", 443, type=socket.SOCK_STREAM)
        dns_blocked = False
        dns_detail = "DNS resolved — out-of-process resolution is NOT blocked"
    except OSError as exc:
        dns_blocked = True
        dns_detail = f"DNS resolution refused ({type(exc).__name__}: {exc})"

    return {
        "blocked": blocked,
        "detail": detail,
        "errno": err,
        "dns_blocked": dns_blocked,
        "dns_detail": dns_detail,
    }


def do_fs_probe(params):
    """Dev-only filesystem-confinement probe (issue #286 PR2d): try to READ the given path and
    report whether the OS denied it. Confined by Landlock, a path outside the worker's allow-set is
    refused with EACCES before any bytes are read; unconfined the read succeeds. Dev-gated
    (PM_SIDECAR_DEV=1) exactly like do_net_selftest, so a shipped worker exposes no
    arbitrary-path-read primitive here at all. Used by the Linux confinement smoke test to prove the
    filesystem restriction is real. The path rides in `probe_path`, NOT `params.path`: the latter is
    the field Rust stages into the granted staging dir, which would defeat the probe."""
    import errno as _errno

    path = params.get("probe_path") or ""
    try:
        with open(path, "rb") as handle:
            handle.read(1)
        return {"denied": False, "detail": "read succeeded — NOT restricted", "errno": None}
    except PermissionError as exc:
        return {"denied": True, "detail": "read refused (restricted)", "errno": exc.errno}
    except OSError as exc:
        denied = exc.errno == _errno.EACCES
        detail = (
            "read refused (filesystem restricted)"
            if denied
            else f"reached the filesystem, NOT restricted ({type(exc).__name__}: {exc})"
        )
        return {"denied": denied, "detail": detail, "errno": exc.errno}


def do_worker_selftest(_params):
    """Confinement preflight (issue #286): exercise the operations the OS sandbox must permit BEFORE
    serving real work, so Rust can fall open (run the worker unconfined) instead of every later
    request failing under a too-tight profile. Three checks, all offline and input-free (no
    untrusted bytes, no network):

      1. `import onnxruntime` — a native `.so`/`.dylib` load that dlopen()s system libraries and
         probes CPU features, exercising the sandbox's executable-map + system-library allow-set
         the way a plain `ping` never would.
      2. list the model-cache dir — a DIRECT read, deliberately NOT routed through the
         offline-masking `_load_model` (which would swallow a sandbox denial into a `ModelNotCached`
         that looks like a cold cache). So a profile that forgot the model-cache grant raises
         `PermissionError` HERE and fails the preflight; an empty dir (a cold cache) lists fine and
         still passes.
      3. when the default model is already cached, embed one string — the ONNX `InferenceSession`
         and the CPU-feature sysctls a bare import never reaches (and it pre-warms the embedder the
         first real request would load anyway). A cold cache raises `ModelNotCached`, which the
         fetch flow handles, so it still passes; only a real sandbox denial makes it raise instead.

    A too-tight sandbox makes one of these raise, main() reports {ok: False}, and the caller falls
    open."""
    import onnxruntime  # noqa: F401 — the import IS a test: a native lib load under the sandbox.

    checked = ["onnxruntime"]
    models_dir = os.environ.get("PM_MODELS_DIR")
    if models_dir and os.path.isdir(models_dir):
        os.listdir(models_dir)  # un-masked: a sandbox denial raises PermissionError here.
        checked.append("models_dir")
    try:
        list(get_embedder().embed(["ok"]))  # consuming the generator is what runs the model.
        checked.append("embed")
    except ModelNotCached:
        checked.append("embed:cold-cache")
    return {"ok": True, "checked": checked, "onnxruntime": onnxruntime.__version__}


HANDLERS = {
    "ping": lambda params: {"ok": True},
    "worker_selftest": do_worker_selftest,
    "convert": do_convert,
    "file_properties": do_file_properties,
    "embed": do_embed,
    "count_tokens": do_count_tokens,
    "rerank": do_rerank,
    "transcribe": do_transcribe,
    "reduce": do_reduce,
    "analyze_image": do_analyze_image,
    "analyze_spreadsheet": do_analyze_spreadsheet,
}

# The probes are unlocked only for debug builds (Rust sets PM_SIDECAR_DEV=1 there, never in
# release), so a shipped worker carries no reachable outbound-socket or arbitrary-path-read
# primitive — the methods are simply unknown and refused. See `SidecarManager::net_selftest`
# (issue #286) and the Linux confinement smoke test.
if os.environ.get("PM_SIDECAR_DEV") == "1":
    HANDLERS["net_selftest"] = do_net_selftest
    HANDLERS["fs_probe"] = do_fs_probe


# Mirror the Rust reader's per-line cap. Rust is the trusted sender and never
# sends anything near this, so a line this large can only be a fault or abuse —
# drop it rather than hand it to json.loads. Env-overridable ONLY so the protocol
# test can exercise the drop path without piping 64 MiB; Rust never sets it, so
# production always runs the full cap.
MAX_LINE_CHARS = int(os.environ.get("PM_SIDECAR_MAX_LINE_CHARS", 64 * 1024 * 1024))


def _quiet_stdout():
    """Send anything a handler prints to stderr instead — stdout is the reply channel.

    One stray `print` inside a third-party library lands mid-line in the newline-delimited JSON, so
    the Rust reader cannot parse the reply, skips the line, and then blocks on an answer that has
    already been sent — wedging the serialized sidecar for the whole per-method timeout. This used
    to be guarded only around the two rapidocr calls, which is the one library we happened to
    catch doing it; MarkItDown, openpyxl, fastembed and faster-whisper all pull in dependency trees
    nobody has audited for prints. Guarding the DISPATCH covers every handler, present and future,
    for the same three lines. `main()` and `fetch_main()` hold the real stdout from before this
    runs, so the reply itself is unaffected."""
    import contextlib

    return contextlib.redirect_stdout(sys.stderr)


def main():
    # The long-lived worker is the process that competes with the user for the machine, so it is
    # the one that steps aside. The --fetch helper is short-lived and network-bound, and lowering
    # its priority would only make a download the user is waiting on feel slower.
    _lower_own_priority()

    # Line-buffered stdout so the Rust side sees each reply immediately. Captured HERE, before any
    # handler runs, because `_quiet_stdout` below swaps `sys.stdout` out from under them — `out`
    # stays the real reply channel.
    out = sys.stdout
    for line in sys.stdin:
        line = line.strip()
        if not line:
            continue
        if len(line) > MAX_LINE_CHARS:
            sys.stderr.write("pm_sidecar: dropping oversized request line\n")
            sys.stderr.flush()
            continue

        req_id = None
        # Bound before the try because the trim check after the reply reads it: a line that fails
        # json.loads never reaches the assignment below, and a stale value from the previous
        # request would otherwise decide this one's housekeeping.
        method = None
        try:
            req = json.loads(line)
            req_id = req.get("id")
            method = req.get("method")
            handler = HANDLERS.get(method)
            if handler is None:
                raise ValueError(f"unknown method: {method!r}")
            with _quiet_stdout():
                result = handler(req.get("params") or {})
            response = {"id": req_id, "ok": True, "result": result}
        except ModelNotCached as miss:
            # Not a failure to report: the offline worker just needs Rust to download this
            # model (via --fetch) and retry. `error_kind` lets request() tell this apart from
            # a genuine error (issue #286). Must precede the broad `except` below — ModelNotCached
            # is an Exception too.
            response = {
                "id": req_id,
                "ok": False,
                "error": str(miss),
                "error_kind": "model_not_cached",
            }
        except Unconvertible as verdict:
            # The engine answered and refused THIS FILE. `error_kind` is the only thing that lets
            # Rust skip the item and let the delta cursor past it — `str(exc)` carries no type, so
            # without this marker a broken engine and a bad file are the same string. Must precede
            # the broad `except` below, which would otherwise swallow it untagged.
            response = {
                "id": req_id,
                "ok": False,
                "error": str(verdict),
                "error_kind": "unconvertible",
            }
        except Exception as exc:  # report, never crash the loop
            traceback.print_exc(file=sys.stderr)
            response = {"id": req_id, "ok": False, "error": str(exc)}

        # `default=str` so a non-JSON-serializable result stringifies instead of raising here
        # (outside the try) and silently killing the read loop.
        #
        # `allow_nan=False` is the load-bearing half. Python's json emits bare `NaN` / `Infinity`
        # for non-finite floats, which are NOT valid JSON: the Rust reader can't parse the line, so
        # it skips it and then blocks forever on a reply that has already been sent. That wedges the
        # whole serialized sidecar -- ingest, chat retrieval, rerank, transcribe, map -- for
        # the full
        # per-method timeout (30 minutes for analyze_image/embed) before the child is killed. One
        # photo with a zero-denominator GPS rational was enough to do it.
        #
        # So a non-finite number becomes a normal, per-request FAILURE: the caller sees an honest
        # error for that one item and the engine keeps serving everything else. `_gps_to_decimal`
        # already refuses to produce one; this is the boundary that guarantees no future
        # handler can.
        try:
            payload = json.dumps(response, default=str, allow_nan=False)
        except ValueError:
            payload = json.dumps(
                {
                    "id": req_id,
                    "ok": False,
                    "error": "non-finite number in result (the value could not be represented)",
                }
            )
        out.write(payload + "\n")
        out.flush()

        # AFTER the reply is on the wire, never before: the caller must not wait on housekeeping.
        # `method` is whatever the request claimed, so an unknown one simply is not in the set.
        if method in _TRIM_AFTER:
            _release_free_memory()


def _ensure_model(method, params):
    """Load — and thus download, with the network allowed — the model `method` needs, so a later
    OFFLINE worker call finds it in the shared cache. Reuses the worker's own lazy loaders; the
    cache location is identical (same OS user, same fastembed/whisper defaults, and whisper's
    `model_dir` rides along in `params`), so what this process fetches is exactly what the worker
    then reads."""
    model = params.get("model") or EMBED_MODEL
    custom = params.get("custom")
    if method in ("embed", "count_tokens"):
        get_embedder(model, custom)
        get_tokenizer(model, custom)
    elif method == "rerank":
        get_reranker(params.get("model"), custom)
    elif method == "transcribe":
        get_whisper(params.get("model_dir"))
    elif method == "analyze_image":
        # Photo OCR. rapidocr fetches its own models from its own host rather than through
        # huggingface_hub, but that is the fetcher's problem, not the worker's — the contract is
        # identical to every other model's, so it rides the same path. Nothing here touches the
        # image: `params` carries no path (Rust strips it) and this only builds the engine.
        get_ocr_engine()
    else:
        raise ValueError(f"nothing to fetch for method {method!r}")


def fetch_main():
    """The network-allowed model DOWNLOADER (issue #286). Rust spawns this SHORT-LIVED process —
    deliberately WITHOUT PM_SIDECAR_OFFLINE — when the offline worker reports a model isn't cached.
    It reads one request line ({"method","params"}, the shape the worker gets), downloads that
    method's model into the shared cache, writes one reply line, and exits. It never parses
    untrusted file bytes: its only input is a trusted model id from Rust, so it is the one component
    that legitimately keeps network access once the worker is sandboxed."""
    # The real reply channel, held before `_quiet_stdout` swaps `sys.stdout` — a downloader that
    # prints its progress (rapidocr's does) would otherwise land on the one line Rust parses.
    out = sys.stdout
    line = sys.stdin.readline()
    try:
        req = json.loads(line)
        with _quiet_stdout():
            _ensure_model(req.get("method"), req.get("params") or {})
        reply = {"ok": True}
    except Exception as exc:
        traceback.print_exc(file=sys.stderr)
        reply = {"ok": False, "error": str(exc)}
    out.write(json.dumps(reply) + "\n")
    out.flush()
    sys.exit(0 if reply.get("ok") else 1)


if __name__ == "__main__":
    if "--fetch" in sys.argv[1:]:
        fetch_main()
    else:
        main()
